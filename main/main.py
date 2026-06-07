import sys

class Logger:
    def __init__(self, filepath):
        self.terminal = sys.stdout
        self.log = open(filepath, "a", encoding="utf-8")
    
    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)
        self.log.flush()
    
    def flush(self):
        self.terminal.flush()
        self.log.flush()
from datetime import datetime
log_filename = f"run_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
sys.stdout = Logger(log_filename)

import os, tempfile
from openpyxl import load_workbook
from llama_cpp import Llama
from concurrent.futures import ThreadPoolExecutor
from LLMFunctions.LlamaResponse import llm_response
from DriveFunctions.Google import Create_Service
from DriveFunctions.FolderDive import get_ParentFolderId, Search_Folder, Search_Folder_with_names
from DriveFunctions.GetImages import get_image_bytes
from OCRFunctions.TextRecognition import Text_from_images, Record_Grouping_with_Dates, new_edit_image
from OCRFunctions.OCRCleaner import master_clean_ocr, remove_sidebar_noise
from OCRFunctions.ICDManual import find_primary_icd, find_other_icd, find_remarks
from ExcelFunctions.OpenExcel import load_in_file, get_column_names
from ExcelFunctions.DataAddition import specific_id_row, row_num_checker, check_row_data, data_per_row, add_data_to_excel
from JSONExtract.JSONtoPy import extract_json, age_checker, extract_last_datetime, pick_later_time,validate_disposition
from JSONExtract.ICDValidator import load_icd_codes, validate_icd, filter_other_icd
from JSONExtract.CheckPointSystem import check_progress, save_progress
os.environ['PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK'] = 'True'
from paddleocr import PaddleOCR
import paddle

CLIENT_FILE = 'credentials.json'
API_NAME = 'drive'
API_VER = 'v3'
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']

filepath = r"g:\ER Visits(Above 60) - Jan 2020 to Jun 2025 - Copy - Copy.xlsx"
icd_valid_codes, icd_valid_prefixes = load_icd_codes(r"g:\icd102019enMeta\icd102019syst_codes.txt")
sheet_name= "Visit Recs"

file = load_in_file(filepath= filepath, sheet_name= sheet_name)
file = file.reset_index(drop=True) 

column_names = get_column_names(file)

wb = load_workbook(filename= filepath)
ws = wb[sheet_name]

checkpointfile = "checkpoint.json"
ALLOWED_IDS = {
    63896, 55084, 140725, 140756, 140885, 134226, 61945, 116807, 141233, 141222,
    141589, 141622, 12087, 142168, 89649, 143153, 32452, 15345, 2845, 143300,
    142563, 111423, 145156, 145193, 145357, 147763, 83710, 75034, 149720, 55803,
    116944, 111457, 161843, 162060, 162244, 162295, 162301, 163362, 163468, 163578,
    144380, 144387, 163349, 162807, 164517, 164636, 164676, 165173, 136019, 165615,
    37222, 165299, 165751, 42936, 165910, 167070, 3562, 167191, 167541, 167950,
    52372, 66543, 167259, 168087, 168117, 168296, 13873, 169096, 169241, 166104,
    143046, 171079, 171503, 172061, 172059, 163286, 133823, 173383, 173389, 173824,
    173978, 174488, 174297, 175208, 175602, 30097, 135628, 157893, 176511, 176530,
    176615, 140201, 177521, 177520, 177522, 177525, 176899, 177593, 16668, 161363,
    177808, 68000, 178105, 2240, 178393, 177024, 178968, 179055, 140199, 179096,
    179097, 142541, 61407, 118036, 119003, 35910, 120262, 120468, 3065, 58148,
    123697, 121510, 123729, 87085, 125069, 126445, 96546, 126985, 56027, 127737,
    128082, 128083, 129042, 948, 130609, 129994, 131439, 131567, 131837, 131907,
    132483, 132552, 76353, 134631, 135244, 135282, 136898, 136505, 137296, 137317,
    137775, 138179, 138551, 138760, 49728, 115847, 2253, 139633, 113646, 132654,
    140005, 140179, 140198, 63914, 140504, 88097, 141012, 141123, 94860, 141220,
    141230, 141238, 141903, 40892, 142693, 142052, 142771, 76289, 144260, 107183,
    17925, 156754, 156878, 156955, 156798, 23514, 22793, 156978, 156980, 156809,
    157124, 157446, 157448, 25447, 151440, 138341, 156841, 158249, 4224, 9465,
    158115, 158814, 159108, 148469, 126629, 12426, 158578, 50532, 160389, 134363,
    159129, 160968, 160895, 116351, 131258, 161625, 161892, 87144, 24, 129081,
    162407, 22164, 162617, 42390, 60384, 162812, 162934, 163004, 163351, 18341,
    163623, 162868, 162852, 73338, 164167, 164177, 164336, 164490, 79464, 164750,
    165024, 50818, 83157, 152929, 165931, 166071, 144473, 165943, 29565, 94165,
    167445, 166577, 167750, 167753, 167301, 167218, 74245, 168142, 167856, 168220,
    128259, 168575, 45252, 60877, 168284, 53310, 169018, 169098, 169308, 156768,
    169446, 50540, 169645, 169651, 169580, 75657, 77383, 169911, 160734, 170209,
    169881, 169925, 170422, 63285, 146883, 171277, 169876, 42298, 140328, 170887,
    165689, 139499, 172644, 172704, 172753, 172893, 166500, 5293, 174000, 174734,
    42070, 174923, 175419, 144385, 145423, 175606, 175619, 55679, 175728, 176112,
    131345, 176219, 176289, 176756, 176766, 176943, 176600, 177124, 103328, 178428,
    178711, 75082, 125645, 125905, 126255, 127663, 129702, 130487, 130959, 132700,
    132526, 133952, 76291, 134962, 135453, 76439, 70807, 136273, 136359, 136354,
    136382, 136490, 136530, 136543, 136619, 87066, 137319, 138146, 138489, 138488,
    139277, 139379, 139380, 55380, 139709, 139949, 140207, 140217, 140361, 140362,
    140485, 140533, 140976, 141562, 42487, 142813, 147873, 147875, 147882, 6277,
    148673, 149382, 149473, 149926, 149930, 149246, 150222, 46415, 150248, 150445,
    150448, 148491, 152369, 152420, 145583, 152990, 128624, 154855, 155050, 155484,
    156391, 156848, 156946, 156951, 156952, 157092, 157131, 157288, 156684, 159358,
    159498, 159650, 160343, 50966, 161074, 161230
}

system_content = '''You are a strict medical data extractor. Extract only 
                  what is explicitly present in the text. Never invent or 
                  infer data. If a field is not found, use null.
                  Always return valid complete JSON. Never use triple quotes.
                  Never leave strings unterminated. Always close all brackets and braces.
                  Do NOT extract from Nurses Notes, Nursing Assessment, Observation Notes, or any reassessment section unless stated otherwise. '''

prompt = '''### GLOBAL RULES:
- Extract from the FIRST occurrence of each field unless stated otherwise.
- Extract only the text directly after the label, stopping at the next section heading or any text ending in ":".
- Do NOT include the label itself in the extracted value.
- Do NOT wrap values in curly braces or any other characters — return plain values only.
- If a field is not found, is empty, or is non-meaningful, return null.

### EXTRACTION RULES:
1. **date_of_birth:** Find "DOB | Age | Gender:". Extract the value before the first vertical bar (|), formatted as DD/MM/YYYY.
2. **nationality:** Value after "Nationality:".
3. **Vitals:** Find the FIRST occurrence of the "Vitals" subheading only. Do NOT mix values from different vitals sections. Find the first row of values under the labels "Temperature", "Pulse", "Respiratory", "BP", "O2SAT". 
    These labels and their values may appear on separate lines.
   - **bp_mmhg:** Find the value under the "BP" label. It MUST contain a forward slash (e.g. "120/80"). If the value does not contain a "/", return null. Strip "mm/Hg". Never return a standalone number as BP.   
   - **temperature_celsius:** Under "Temperature". Strip °C or *C, return numeric only.
   - **pulse_min:** Under "Pulse". Strip "/min", return numeric only.
   - **respiratory_min:** Under "Respiratory". Strip "/min", return numeric only.
   - **o2_sat:** Under "O2SAT". If no number is directly present, find the first numeric value near a /% symbol in the Nursing Assessment section. Return numeric only.
4. **visit_date / visit_time:** Find "Visit Date:" followed by DD/MM/YYYY then a 4-digit time. Split into separate fields. If the date separator is missing (e.g. "08/08 2020"), reconstruct as DD/MM/YYYY.
5. **disposition_date / disposition_time:** Find the table containing the "Disposition Done" column at the END of the record. Extract ONLY the date and time from that column. This is NOT the visit time, triage time, or any timestamp from Nursing Assessment or Observation Notes.
6. **pain_scale_score:** Find "Numerical(X)". Return only the integer X. Do not confuse with GCS.
7. **gcs:** Find the GCS value in the "Nursing Assessment" vitals section, typically "XX/15". Return only the numerator as a plain integer.
8. **triage_category:** Return only the number after "Triage Category:".
9. **past_history:** Find "Past History:" or "Past Medical History". Stop before "Travel History:" or any sidebar navigation text such as "HEMODIALYSIS", "LAB REPORTS", "OTHER DOCUMENTS", "ASSESSMENT/RE-ASSESSMENT". Do NOT include any section headings or subheadings in the extracted value.
Strip any leading newlines or whitespace.
10. **occupation:** Value after "Occupation:".
11. **marital_status:** Value after "Marital Status:".
12. **advice_health_education:** Find text from "Advice & Health Education:". Stop at "Education Given To:".
13. **condition_at_disposition:** Value after "Condition at the time of Disposition:" or "Condition at the tine of Disposition:".
14. **disposition_type:** Value after "Disposition Type:". If no genuine free-text clinical note exists at the end of that section, return null.
15. **travel_history:** If the value is empty, a date, month-year, or any non-place text, return null. If "Travel History:" is immediately followed by another label, return null.
16. **current_medication:** Find "Current Medication". Stop at "Medical Prescription" or "Medication Order". If the section contains "No Important History" or is empty, return null. Do NOT extract from Medication Order, Medical Prescription, Injection Administration Log, or Care Plan.
17. **psychosocial:** Value after "Psychosocial:". Do not include occupation.
18. **disease_grouping:** Leave as null unless specified.
'''

paddle.set_flags({
"FLAGS_fraction_of_cpu_memory_to_use": 1.0,
"FLAGS_allocator_strategy": "naive_best_fit",
})

ocr_engine = PaddleOCR(
    use_doc_orientation_classify=False,
    use_doc_unwarping=False,
    use_textline_orientation= True,
    text_recognition_batch_size=2, 
    text_det_limit_side_len= 2000,
    text_det_limit_type= 'max',
    text_det_box_thresh= 0.5,
    text_det_thresh= 0.5,
    text_det_unclip_ratio=1.4, #Best results with 1.6
    lang='en',
    device= 'cpu',
    enable_mkldnn=True,
    mkldnn_cache_capacity=10,
    cpu_threads=4
    )

llm_engine = Llama(model_path= r"g:\models\qwen2.5-7b-instruct-q5_0-00001-of-00002.gguf",
                chat_format= "chatml",
                flash_attn= True,
                n_gpu_layers=-1,
                temperature= 0.2,
                seed= 1337,
                n_ctx= 8192,
                verbose= False)

service = Create_Service(CLIENT_FILE, API_NAME, API_VER, SCOPES)

parent_folder_id = get_ParentFolderId(body= service, 
                                      text= 'Elderly ER Data Pics')

subfolder_map = Search_Folder_with_names(body=service, parent_id=parent_folder_id)
subfolder_ids = sorted(subfolder_map.keys(), key=lambda fid: int(subfolder_map[fid].strip()))

checkpoint = check_progress(checkpointfile)

for folder in subfolder_ids:
    llm_engine.reset()
    comp_id = subfolder_map[folder].strip()
    print(comp_id)
    if int(comp_id) not in ALLOWED_IDS:
        continue
    if int(comp_id) in checkpoint:
        print("Skipped")
        continue

    rows= specific_id_row(dataframe= file, specific_id= comp_id)
    image_ids = Search_Folder(body= service, parent_id= folder)
    image_bytes, pdf_bytes = get_image_bytes(image_file_ids= image_ids, service= service)

    texts = []
    readable_img = []
    pdf_texts = []
    if image_bytes:
        with ThreadPoolExecutor(max_workers=16) as Executor:
            if image_bytes:
                readable_img = list(Executor.map(new_edit_image, image_bytes))
            texts = list(Text_from_images(ocr= ocr_engine, readable_list= readable_img))

    elif pdf_bytes:
        for pdf_data in pdf_bytes:
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete = False) as tmp:
                tmp.write(pdf_data)
                tmp_path = tmp.name
            try:
                result = ocr_engine.predict(tmp_path)
                for item in result:
                    texts_pdf = item.get('rec_texts', [])
                    pdf_texts.append("\n".join(texts_pdf))
            finally:
                os.unlink(tmp_path)
        texts = pdf_texts
    else:
        texts = []
        with open("empty_folders.txt", "a") as f:
            f.write(f"{comp_id}\n")
        print(f"No files found for comp {comp_id}")
        continue
    print("OCR done!")

    record = Record_Grouping_with_Dates(texts= texts)
    record = {key: value for key, value in record.items() if key != 'Unknown' or value}

    row_data_check = check_row_data(rows= rows)
    row_count_check = row_num_checker(rows=row_data_check, total_records=record)
    if not row_data_check:
        continue
    copy_row_data = None

    if row_count_check > 0:
        new_row = max(row_data_check) + 3
        ws.insert_rows(new_row, row_count_check)
        copy_row_data = [cell.value for cell in ws[(max(row_data_check)+2)]]
        current_max = max(row_data_check)
        for j in range(1, row_count_check + 1):
            row_data_check.append(current_max + j)
        wb.save(filename=filepath)
        file = load_in_file(filepath=filepath, sheet_name=sheet_name)
        file = file.reset_index(drop=True)
        column_names = get_column_names(file)
        rows = specific_id_row(dataframe=file, specific_id=comp_id)

    new_record = data_per_row(records= record, df= file, comp_id= comp_id)
    if not new_record:
        continue
    print("Cleaning Started!")
    clean_records = master_clean_ocr(records_dict= new_record)
    clean_noise_records = remove_sidebar_noise(records= clean_records)
    print(f"Records expected: {len(new_record)}, clean_noise_records: {len(clean_noise_records)}")

    primary_icd_list = list(find_primary_icd(clean_noise_records))
    other_icd_list = list(find_other_icd(clean_noise_records, primary_icd_list))
    remarks_list = list(find_remarks(clean_noise_records))

    print("LLM processing...")
    llm_step = list(llm_response(llm= llm_engine, 
                                 clean_records= clean_noise_records, 
                                 column_names= column_names, 
                                 prompt= prompt, 
                                 system_content= system_content, 
                                 primary_icd= primary_icd_list, 
                                 other_icd_list= other_icd_list, 
                                 remarks_list= remarks_list
                                 ))
    for i,record in enumerate(llm_step):
        data = extract_json(record)
        if data is None:
            print(f"Skipping record {i+1} for comp {comp_id} - failed to parse JSON")
            continue
        if i >= len(row_data_check):
            print(f"Skipping record {i+1} for comp {comp_id} - no row index available")
            continue
        regex_icd = primary_icd_list[i] if i < len(primary_icd_list) else None
        if regex_icd and data[22]:
            primary_code = data[22].split(' - ')[0].strip() if ' - ' in str(data[22]) else str(data[22])
            if regex_icd != primary_code:
                print(f"ICD mismatch comp {comp_id} record {i+1}: regex={regex_icd}, llm={primary_code}")
                with open("icd_mismatches.txt", "a") as f:
                    f.write(f"comp {comp_id} record {i+1}: regex={regex_icd}, llm={primary_code} — LLM OVERRODE REGEX\n")
        if data[23]:
            data[23] = filter_other_icd(data[23], icd_valid_codes, icd_valid_prefixes, comp_id, i+1)
        if data[22]:
            primary_code = data[22].split(' - ')[0].strip() if ' - ' in str(data[22]) else str(data[22])
            if not validate_icd(primary_code, icd_valid_codes, icd_valid_prefixes):
                print(f"Invalid primary ICD for comp {comp_id} record {i+1}: {primary_code}")
                with open("invalid_icds.txt", "a") as f:
                    f.write(f"comp {comp_id} record {i+1}: invalid primary ICD flagged: {primary_code}\n")
                data[22] = None

        disp_date, disp_time = extract_last_datetime(clean_noise_records[i])
        print(f"Record {i+1} disposition regex result: {disp_date}, {disp_time}")
        
        final_date, final_time = pick_later_time(disp_date, disp_time, data[29], data[30])
        data[29] = final_date
        data[30] = final_time
        data = validate_disposition(data, comp_id= comp_id, record_num= i+1)

        calculated_age = age_checker(age=None, dob=data[2], visit=data[3])
        if calculated_age:
            ws.cell(row=row_data_check[i] + 2, column=5, value=calculated_age)
        print(f"\n\nLLM Output for record {i+1}:\n{data}\n\n")
        add_data_to_excel(data= data, 
                          ws= ws, 
                          starting_column= 6, 
                          row_num= row_data_check[i], 
                          initial_data= copy_row_data 
                          if row_count_check > 0 else None)
    save_progress(checkpointfile, comp_id= comp_id)
    checkpoint.add(int(comp_id))
    wb.save(filename= filepath)